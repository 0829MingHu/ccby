import requests
import openpyxl
from time import sleep
import random

wb=openpyxl.load_workbook(r'youtube.xlsx')
sheet=wb.active
for i in range(1,sheet.max_row+1):
    sleep(1 + random.random())
    url='https://www.youtube.com/watch?v={}'.format(sheet.cell(row=i,column=1).value)
    print(i)
    print(url)
    res=requests.get(url).text
    if 'Creative Commons Attribution license' in res:
        sheet.cell(row=i,column=2).value='true'
        print('true')
    else:
        sheet.cell(row=i,column=2).value='false'
        print('false')
    if i % 10 == 0:
        wb.save('a.xlsx')